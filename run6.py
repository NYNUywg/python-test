import json
import os
import time
from datetime import datetime

import requests
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from playwright.async_api import async_playwright
from pyppeteer import launch
import asyncio


def append_to_excel(data, filename):
    directory = 'data'
    if not os.path.exists(directory):
        os.makedirs(directory)
    filename = os.path.join(directory, filename)
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    # for row in data:
    ws.append(data)

    wb.save(filename)


def post(url, data):
    json_data = json.dumps(data)
    headers = {
        'Content-Type': 'application/json',
        'Cookie': 'HWWAFSESID=090d6dcbb843b5c5e6; HWWAFSESTIME=1723626452297'
    }
    # 发送POST请求
    response = requests.post(url, headers=headers, data=json_data)
    return response


# 获取所有的公司uid
def get_all_uid():
    country_id = 78
    uid_list = []
    country_name = "Kuwait"
    total = 243

    # 读取文件
    workbook_path = f"./data/{country_name}_{country_id}_{total}.xlsx"
    if os.path.exists(workbook_path):
        workbook = load_workbook(workbook_path)
        sheet = workbook.active
        start_line_number = sheet.max_row + 1
    else:
        start_line_number = 1
    with open(f"./country/{country_name}_{country_id}_{total}.txt", 'r') as file:
        lines = file.readlines()
        for line_number, line in enumerate(lines, start=1):
            if line_number < start_line_number:
                continue
            uid_list.append(line.strip())
    return country_name, country_id, uid_list, total


async def get_index_context(page, url):
    company = ""
    email = ""
    phone = ""
    await page.goto(url)
    await page.wait_for_load_state('networkidle')

    h1_elements = await page.locator('xpath=//h1').all()
    if len(h1_elements) >= 2:
        company = await h1_elements[2].inner_text()

    content_elements = await page.locator('.content').all()
    if len(content_elements) >= 3:
        email = await content_elements[-3].inner_text()
    if len(content_elements) >= 2:
        phone = await content_elements[-2].inner_text()
    if "@" in email:
        pass
    else:
        email = phone
        phone = ""
    return company, email, phone


async def fetch_data_with_retry(page, url, retry=10):
    for _ in range(retry):
        try:
            company, email, phone = await get_index_context(page, url)
            return company, email, phone
        except Exception as e:
            print(f"Failed to fetch data from {url}. Retrying...")
    return None, None, None


async def main():
    cookie_value = "19ca1ae55aab4903806925a7d71ae4aa"
    country_name, country_id, uid_list, total = get_all_uid()

    async with async_playwright() as p:
        browser = await p.chromium.launch()
        context = await browser.new_context()
        page = await context.new_page()
        await context.add_cookies([{
            "domain": ".jctrans.com",
            "name": "JC-JAVA-Token",
            "path": "/",
            "value": cookie_value
        }])

        count = 0
        for uid in uid_list:
            url = 'https://www.jctrans.com/cn/home/' + uid
            company, email, phone = await fetch_data_with_retry(page, url)
            count += 1
            print(count, company, email, phone)
            filename = f"{country_name}_{country_id}_{total}.xlsx"
            append_to_excel([company, email, phone], filename)
        await browser.close()


if __name__ == '__main__':
    import asyncio

    asyncio.run(main())
